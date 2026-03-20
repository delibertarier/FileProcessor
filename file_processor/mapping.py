from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Optional

from openpyxl import load_workbook


TransformFunc = Callable[[str], str]


@dataclass
class MappingEntry:
    """Single mapping from a source CSV column to an XML XPath."""

    source_column_index: int  # 1-based CSV column index
    target_xpath: str
    transform_name: str | None = None


@dataclass
class Mapping:
    """In-memory representation of all mappings from XLS."""

    entries: list[MappingEntry]


def load_mapping_from_xls(path: Path) -> Mapping:
    """
    Load mapping definition from an XLS file.

    Convention (first sheet, from row 2 onward):
    - Column 1: source column index (1-based).
    - Column 2: target XPath.
    - Columns 3-5: free for codes/remarks (ignored by the loader).
    - Column 6 (F): optional transform rule name (e.g. `first_3`, `skip`).
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]

    # Read all relevant rows first so we can detect 0-based vs 1-based column numbering.
    raw_rows = list(ws.iter_rows(min_row=2, values_only=True))

    # Detect whether the mapping uses 0-based column numbers:
    # if the first non-empty numeric source index is 0, we treat the whole sheet as 0-based
    # and add +1 to all column numbers.
    offset = 0
    for row in raw_rows:
        source_idx_raw = (row + (None,))[0]
        if source_idx_raw in (None, ""):
            continue
        try:
            first_index = int(source_idx_raw)
        except (TypeError, ValueError):
            continue
        if first_index == 0:
            offset = 1
        break

    entries: list[MappingEntry] = []

    for row in raw_rows:
        # We only care about:
        # - Column 1: source index
        # - Column 2: XPath
        # - Column 6 (F): transform name
        row_padded = (row + (None,) * 6)[:6]
        source_idx_raw = row_padded[0]
        xpath_raw = row_padded[1]
        transform_raw = row_padded[5]
        if xpath_raw is None:
            # Skip empty rows
            continue

        # We only support mappings from actual CSV columns.
        if source_idx_raw in (None, ""):
            continue
        try:
            source_index = int(source_idx_raw) + offset
        except (TypeError, ValueError):
            raise ValueError(f"Invalid source column index in mapping XLS: {source_idx_raw!r}")

        entry = MappingEntry(
            source_column_index=source_index,
            target_xpath=str(xpath_raw).strip(),
            transform_name=(str(transform_raw).strip() if transform_raw not in (None, "") else None),
        )
        entries.append(entry)

    return Mapping(entries=entries)

